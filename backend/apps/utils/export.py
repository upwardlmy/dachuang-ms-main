
import os
import zipfile
from io import BytesIO
from datetime import datetime
import openpyxl  # type: ignore[import-untyped]
from django.utils.timezone import localtime

def generate_excel(data, headers, filename="export.xlsx"):
    """
    Generate an Excel file from a list of dictionaries.
    :param data: List of dictionaries or objects
    :param headers: Dictionary mapping field names to column headers {field: header}
    :param filename: Filename for the download
    :return: BytesIO object containing the Excel file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export Data"

    # Write headers
    header_keys = list(headers.keys())
    for col_idx, header in enumerate(headers.values(), 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data
    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(header_keys, 1):
            value = getattr(item, field, None) if hasattr(item, field) else item.get(field)
            
            # Handle Choice Fields (get display value)
            if hasattr(item, f"get_{field}_display"):
                 value = getattr(item, f"get_{field}_display")()
            
            # Handle standard fields
            if isinstance(value, datetime):
                value = localtime(value).strftime("%Y-%m-%d %H:%M:%S")
            elif value is None:
                value = ""
            
            ws.cell(row=row_idx, column=col_idx, value=str(value))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_zip(files, filename="attachments.zip"):
    """
    Generate a ZIP file from a list of file paths.
    :param files: List of tuples (file_path_on_disk, arcname_in_zip)
    :param filename: Filename for the download
    :return: BytesIO object containing the ZIP file
    """
    output = BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path, arcname in files:
            if os.path.exists(file_path):
                zf.write(file_path, arcname)
    
    output.seek(0)
    return output
